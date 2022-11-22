# Imports the required modules
import openpyxl


def fastq_reader_no_spike(fastq_path, umi_length, homopolymer_length):
    """
    Reads and saves the info from the .fastq file to a list.
    :param fastq_path: Path to the .fastq file.
    :param umi_length: Length of the unique molecular identifier. Entered by
            user in settings.py.
    :param homopolymer_length: Maximum length of homopolymers within barcodes.
    :return fastq_data: List with the structure [barcode 1, barcode2, etc].
    """
    # Creates an empty list, dictionary and set
    fastq_data = []
    homopolymers = {"A": {}, "T": {}, "C": {}, "G": {}}
    umi_set = set()

    try:
        # Opens the .fastq and reads it line by line
        with open(fastq_path, "r") as fastq_file:
            # Checks if the user used UMIs
            if umi_length == 0:
                for line in fastq_file:
                    barcode = line.split(":")[-1].replace("\n", "")
                    # If line starts with a @ it retrieves the barcode
                    # ands saves it in a list
                    if line.startswith("@") and "N" not in barcode:
                        fastq_data.append(barcode)

                        # Checks if user would like to analyse
                        # homopolymers
                        if homopolymer_length > 0:
                            homopolymers = homopolymer_check(
                                homopolymer_length, barcode, homopolymers)

            else:
                for line in fastq_file:
                    # If line starts with a @ it retrieves the barcode
                    if line.startswith("@") and "N" not in line.split(":")[-1]:
                        # Retrieves the UMI sequence from the header
                        umi = line.replace("\n", "").split(":")[-1].\
                            split("+")[0][-umi_length:]
                        barcode = line.replace("\n", "").split(":")[-1].\
                            replace(umi+"+", "+")
                        # Adds the barcodes to a list if UMI has not
                        # been found in previous reads
                        if umi not in umi_set:
                            umi_set.add(umi)
                            fastq_data.append(barcode)

                        # Checks if user would like to analyse
                        # homopolymers
                        if homopolymer_length > 0:
                            homopolymers = homopolymer_check(
                                homopolymer_length, barcode, homopolymers)

        if len(fastq_data) == 0:
            print("Error 15: Fasta file format is incorrect. No headers "
                  "found.")
            exit(14)

        # Returns the fastq list and homopolymers dictionary
        return fastq_data, homopolymers

    except FileNotFoundError:
        print("Error 16: Entered fastq file has not been found.")
        exit(15)


def barcode_file_reader(barcode_file, sequencing_method, analyse_combination):
    """
    Retrieves the barcodes (and spike in sequences if spike_in parameter is
    '2') from the entered Excel file.
    :param barcode_file: Path to the Excel file.
    :param sequencing_method: Parameter from settings.py.
    :param analyse_combination: Parameter from settings.py.
    :return barcode_file_list: List with all barcodes from the Excel file.
    :return barcode_file_dict: Dictionary with structure {barcode: [well, spike
            seq, 0]}.
    """
    try:
        # Loads in the Excel barcode file
        excel_reader = openpyxl.load_workbook(barcode_file)
        sheet = excel_reader.active

    except FileNotFoundError:
        # Returns an error if file has not been found
        print("Error 17: Entered barcode file has not been found.")
        exit(16)

    # Creates an empty list and dictionary
    barcode_file_list = []
    barcode_file_dict = {}

    # Checks sequencing method to retrieve correct i5 column
    if sequencing_method == 1:
        i5 = 8
    else:
        i5 = 9
    try:
        # Checks spike in parameter
        if analyse_combination == 1:
            # Retrieve barcodes from the Excel file
            for row in sheet.iter_rows(min_row=4, values_only=True):
                barcode_file_list.append([row[0], row[4] + "+" + row[i5]])
            # Returns the barcode_file_list
            return barcode_file_list

        else:
            # Retrieve barcodes and spike-in sequences from the Excel file
            for row in sheet.iter_rows(min_row=4, values_only=True):
                barcode = row[4] + "+" + row[i5]
                barcode_file_dict[barcode] = [row[0], row[-1], 0]
            # Returns the barcode_file_dict
            return barcode_file_dict

    # Returns an error if file format is incorrect
    except IndexError or TypeError:
        print("Error 18: Barcode file format incorrect.")
        exit(17)


def fastq_reader_with_spike(fastq_path, trim_left, trim_right, umi_length,
                            homopolymer_length):
    """
    Reads and saves the info from the .fastq file to a dictionary.
    :param fastq_path: Path to the .fastq file.
    :param trim_left: Parameter to trim the left side of the spike sequence.
    :param trim_right: Parameter to trim the right side of the spike sequence.
    :param umi_length: Length of the unique molecular identifier. Entered by
            user in settings.py.
    :param homopolymer_length: Maximum length of homopolymers within barcodes.
    :return fastq_data: List with the structure [[i5, i7, sequence], [i5, i7,
            sequence], etc].
    """
    # Creates an empty list, dictionary and set
    fastq_data = []
    homopolymers = {"A": {}, "T": {}, "C": {}, "G": {}}
    umi_set = set()

    try:
        # Opens the .fastq and reads it line by line
        with open(fastq_path, "r") as fastq_file:
            # Checks if the user used UMIs
            if umi_length == 0:
                for line in fastq_file:
                    barcode = line.split(":")[-1].replace("\n", "")
                    # If line starts with a @ it retrieves the barcode
                    if line.startswith("@") and "N" not in barcode:
                        templist = barcode.split("+")

                        # Checks if user would like to analyse
                        # homopolymers
                        if homopolymer_length > 0:
                            homopolymers = homopolymer_check(
                                homopolymer_length, barcode, homopolymers)

                    # Checks if line contains only letters and saves it to a
                    # variable
                    elif line[:-1].isalpha() and len(templist) > 0:

                        # Saves the sequence to list
                        templist.append(line[trim_left:-trim_right-1])
                        fastq_data.append(templist)

            else:
                umi_bool = False
                for line in fastq_file:
                    # If line starts with a @ it retrieves the barcode
                    if line.startswith("@") and "N" not in line.split(":")[-1]\
                            .replace("\n", ""):
                        templist = line.split(":")[-1].replace("\n", "")\
                            .split("+")
                        # Checks if UMI has been found in previous reads
                        if templist[0][-umi_length:] in umi_set:
                            umi_bool = True
                        else:
                            umi_bool = False
                            # Adds the UMI sequence to a set
                            umi_set.add(templist[0][-umi_length:])
                            templist[0] = templist[0][:-umi_length]
                    elif not umi_bool:
                        # Checks if line contains only letters and saves it to
                        # a variable
                        if line[:-1].isalpha() and len(templist) > 0:

                            # Saves the sequence to list
                            templist.append(line[trim_left:-trim_right - 1])
                            fastq_data.append(templist)

                            # Checks if user would like to analyse
                            # homopolymers
                            if homopolymer_length > 0:
                                homopolymers = homopolymer_check(
                                    homopolymer_length, templist[0]+"+"+templist[1], homopolymers)

        # Checks if fastq data is empty and returns an error if empty
        if len(fastq_data) == 0:
            print("Error 19: Fasta file format is incorrect. No headers or "
                  "sequences found.")
            exit(18)

        # Returns the fastq list and homopolymers dictionary
        return fastq_data, homopolymers

    except FileNotFoundError:
        # Returns an error if fastq file has not been found
        print("Error 20: Entered fastq file has not been found.")
        exit(19)


def homopolymer_check(homopolymer_length, barcode, homopolymers):
    """
    Checks if the barcodes contain homopolymers longer than the allowed
    homopolymer length.
    :param homopolymer_length: Maximum length of homopolymers within barcodes.
    :param barcode: Barcode from a specific fastq read.
    :param homopolymers: Dictionary containing all found homopolymers. Has the
        structure {"A": {barcode 1: count, barcode 2: count}, "T": {barcode 1:
        count, barcode 2: count}, etc}.
    :return: Updated dictionary containing all found homopolymers. Has the
        structure {"A": {barcode 1: count, barcode 2: count}, "T": {barcode 1:
        count, barcode 2: count}, etc}.
    """
    # Creates a list
    nucleotide_list = ["A", "T", "C", "G"]

    # Loops through the nucleotide list
    for nucleotide in nucleotide_list:
        # Checks if barcode sequence contains a homopolymer
        if nucleotide * (homopolymer_length + 1) in barcode:
            # Saves the barcode to the homopolymer dictionary
            if barcode in homopolymers[nucleotide]:
                homopolymers[nucleotide][barcode] += 1
            else:
                homopolymers[nucleotide][barcode] = 1

    # Returns the homopolymer dictionary to the filereader functions
    return homopolymers
